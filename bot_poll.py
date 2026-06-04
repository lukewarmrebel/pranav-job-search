"""
Telegram bot poller: processes button callbacks, job URL messages, and slash commands.
Run every 15 minutes via GitHub Actions.

Commands:
  /apply   — generate and send final apply-list Excel (shortlisted + manual)
  /status  — show shortlist and manual counts
  /clear   — reset today's shortlist and manual list
  /help    — show usage
"""
import json, os, sys, uuid, urllib.request, urllib.error
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

TOKEN   = os.environ.get("TELEGRAM_TOKEN", "")
CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "")

if not TOKEN or not CHAT_ID:
    print("No credentials — exiting.")
    sys.exit(0)

Path("data").mkdir(exist_ok=True)

OFFSET_FILE = "data/update_offset.txt"
TODAY_JOBS  = "data/today_jobs.json"
SHORTLIST   = "data/shortlisted.json"
MANUAL      = "data/manual_jobs.json"
SENT_MSGS   = "data/sent_messages.json"


# ── Telegram API helpers ──────────────────────────────────────────────────────

def tg_api(method, payload):
    url  = f"https://api.telegram.org/bot{TOKEN}/{method}"
    data = json.dumps(payload).encode()
    req  = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        print(f"WARNING: {method} HTTP {e.code}: {e.read().decode(errors='replace')}")
        return {}
    except Exception as e:
        print(f"WARNING: {method} failed: {e}")
        return {}

def send_text(text, chat_id=None):
    return tg_api("sendMessage", {
        "chat_id": chat_id or CHAT_ID, "text": text,
        "parse_mode": "HTML", "disable_web_page_preview": True,
    })

def answer_callback(cq_id, text="", alert=False):
    tg_api("answerCallbackQuery", {
        "callback_query_id": cq_id, "text": text, "show_alert": alert,
    })

def edit_text(message_id, text):
    tg_api("editMessageText", {
        "chat_id": CHAT_ID, "message_id": message_id,
        "text": text, "parse_mode": "HTML",
        "disable_web_page_preview": True,
        "reply_markup": {"inline_keyboard": []},
    })

def remove_keyboard(message_id):
    tg_api("editMessageReplyMarkup", {
        "chat_id": CHAT_ID, "message_id": message_id,
        "reply_markup": {"inline_keyboard": []},
    })

def send_document(path, caption=""):
    boundary = uuid.uuid4().hex
    filename = os.path.basename(path)
    with open(path, "rb") as f:
        file_bytes = f.read()
    body = (
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"chat_id\"\r\n\r\n{CHAT_ID}\r\n"
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"caption\"\r\n\r\n{caption}\r\n"
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"document\"; filename=\"{filename}\"\r\n"
        f"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
    ).encode() + file_bytes + f"\r\n--{boundary}--\r\n".encode()
    url = f"https://api.telegram.org/bot{TOKEN}/sendDocument"
    req = urllib.request.Request(url, data=body,
                                  headers={"Content-Type": f"multipart/form-data; boundary={boundary}"})
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        print(f"WARNING: sendDocument HTTP {e.code}: {e.read().decode(errors='replace')}")
        return {}
    except Exception as e:
        print(f"WARNING: sendDocument failed: {e}")
        return {}


# ── state helpers ─────────────────────────────────────────────────────────────

def load_json(path, default):
    p = Path(path)
    if p.exists():
        try: return json.loads(p.read_text())
        except: pass
    return default

def save_json(path, data):
    with open(path, "w") as f:
        json.dump(data, f, indent=2)

# ── scoring (mirrors telegram_notify.py) — used as fallback ──────────────────

def _role_score(t):
    t = t.lower()
    for kw in ["principal product manager","senior product manager","lead product manager",
                "staff product manager","product manager, senior"]:
        if kw in t: return 40, "Product Manager"
    if "product manager" in t:   return 40, "Product Manager"
    if "product owner" in t:     return 37, "Product Owner"
    for kw in ["senior business analyst","lead business analyst","technical business analyst"]:
        if kw in t: return 35, "Business Analyst"
    if "business analyst" in t:  return 35, "Business Analyst"
    for kw in ["consultant, strategy","strategy, growth","transformation consultant",
                "management consultant","digital transformation","strategy consultant"]:
        if kw in t: return 33, "Consultant/Strategy"
    if "consultant" in t:        return 33, "Consultant/Strategy"
    for kw in ["program manager","delivery manager","project manager"]:
        if kw in t: return 28, "Program Manager"
    for kw in ["product specialist","product analyst","product associate","product operations"]:
        if kw in t: return 28, "Product Manager"
    return 0, None

def _skills(d):
    if not d: return 0
    d = d.lower(); p = 0
    if "roadmap" in d or "product strategy" in d: p += 5
    if any(x in d for x in ["backlog","sprint","agile","scrum"]): p += 4
    if any(x in d for x in ["requirements","acceptance criteria","uat"]): p += 4
    if "stakeholder" in d: p += 4
    if any(x in d for x in ["sql","power bi","kpi","analytics"]): p += 4
    if any(x in d for x in ["generative ai","machine learning","llm","agentic","ai/ml","ai product","ai-powered"]): p += 5
    if any(x in d for x in ["digital transformation","process improvement","automation"]): p += 4
    return min(p, 30)

def _domain(d):
    if not d: return 0
    d = d.lower(); p = 0
    if any(x in d for x in ["insurance","motor claims","claims"]): p += 10
    elif any(x in d for x in ["bfsi","fintech","financial services","banking","payments"]): p += 7
    elif any(x in d for x in ["regtech","risk management","compliance"]): p += 4
    if any(x in d for x in ["mba or relevant advanced degree is a must","mba or another advanced degree","mba preferred"]): p += 5
    if any(x in d for x in ["ai product","generative ai","agentic","ai platform","ai strategy"]): p += 5
    return min(p, 20)

def _brand(c, d=""):
    c = (c or "").lower(); d = (d or "").lower()
    if any(x in c for x in ["amazon","google","microsoft","salesforce","jpmorgan","goldman","wells fargo","blackrock"]): return 10
    if "hackajob" in c and "jp morgan" in d: return 10
    if any(x in c for x in ["deloitte","pwc","hsbc","synchrony","nielsen","experian","optum","simcorp","natwest","state street","broadridge","wise"]): return 7
    if any(x in c for x in ["wipro","infosys","tcs","genpact","capgemini","accenture","cognizant","publicis sapient","endava"]): return 4
    return 0

def _salary(j):
    lo = j.get("min_amount"); hi = j.get("max_amount"); cur = j.get("currency") or ""
    if lo and hi: return f"{cur}{int(lo):,}–{int(hi):,}"
    if lo:        return f"{cur}{int(lo):,}+"
    return ""

def build_jobs_from_raw(raw_path="jobs_raw.json"):
    try:
        with open(raw_path) as f:
            raw = json.load(f)
    except Exception:
        return str(date.today()), {}
    run_date_str = raw.get("run_date", str(date.today()))
    from datetime import timedelta
    try:
        run_date = date.fromisoformat(run_date_str)
    except ValueError:
        run_date = date.today()
    cutoff = (run_date - timedelta(hours=48)).isoformat()
    scored = []; seen = set()
    for j in raw.get("jobs", []):
        dp = j.get("date_posted"); title = j.get("title","") or ""; company = j.get("company","") or ""
        desc = j.get("description","") or ""
        if dp and dp < cutoff: continue
        r1, rt = _role_score(title)
        if r1 == 0: continue
        key = (company.lower()[:20], title.lower()[:35])
        if key in seen: continue
        seen.add(key)
        total = r1 + _skills(desc) + _domain(desc) + _brand(company, desc)
        if total < 55: continue
        scored.append({"score":total,"title":title,"company":company,"url":j.get("job_url","") or "",
                        "date":dp or "Recent","role":rt or "","location":j.get("location","") or "",
                        "salary":_salary(j),"site":j.get("site","") or ""})
    scored.sort(key=lambda x: -x["score"])
    for i, j in enumerate(scored): j["idx"] = i
    return run_date_str, {j["idx"]: j for j in scored}


# ── load jobs index — today_jobs.json first, fall back to jobs_raw.json ───────

today_data   = load_json(TODAY_JOBS, {})
session_date = today_data.get("date", str(date.today()))
jobs_by_idx  = {j["idx"]: j for j in today_data.get("jobs", [])}

if not jobs_by_idx:
    print("today_jobs.json empty or missing — rebuilding from jobs_raw.json")
    session_date, jobs_by_idx = build_jobs_from_raw()
    # Persist so future runs don't need to rebuild
    if jobs_by_idx:
        Path("data").mkdir(exist_ok=True)
        with open(TODAY_JOBS, "w") as f:
            json.dump({"date": session_date, "jobs": list(jobs_by_idx.values())}, f, indent=2)


def get_shortlist():
    sl = load_json(SHORTLIST, {"date": session_date, "jobs": []})
    if sl.get("date") != session_date:
        sl = {"date": session_date, "jobs": []}
    return sl

def get_manual():
    m = load_json(MANUAL, {"date": session_date, "entries": []})
    if m.get("date") != session_date:
        m = {"date": session_date, "entries": []}
    return m


# ── apply-list Excel ──────────────────────────────────────────────────────────

def make_apply_excel(shortlisted, manual_entries, date_str):
    if not HAS_OPENPYXL:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Apply List"

    HDR_FILL    = PatternFill("solid", fgColor="1F4E79")
    SCRAPED_FILL= PatternFill("solid", fgColor="E2EFDA")
    MANUAL_FILL = PatternFill("solid", fgColor="DDEEFF")
    HDR_FONT    = Font(bold=True, color="FFFFFF", size=10)
    BD          = Side(style="thin", color="CCCCCC")
    BORDER      = Border(left=BD, right=BD, top=BD, bottom=BD)
    URL_FONT    = Font(color="0563C1", underline="single")

    cols = ["#", "Source", "Score", "Role", "Title", "Company",
            "Location", "Date", "Salary", "Apply URL", "Notes"]
    ws.append(cols)
    for col in range(1, len(cols) + 1):
        c = ws.cell(1, col)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER

    rank = 0
    for job in shortlisted:
        rank += 1
        row = [rank, "Scraped", job.get("score", ""), job.get("role", ""),
               job.get("title", ""), job.get("company", ""), job.get("location", ""),
               job.get("date", ""), job.get("salary", ""), job.get("url", ""), ""]
        ws.append(row)
        for col in range(1, len(row) + 1):
            c = ws.cell(ws.max_row, col)
            c.fill = SCRAPED_FILL; c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
        if job.get("url"):
            c = ws.cell(ws.max_row, 10)
            c.hyperlink = job["url"]; c.font = URL_FONT

    for entry in manual_entries:
        rank += 1
        added = entry.get("added_at", "")[:10]
        url   = entry.get("url", "")
        title = entry.get("title", "")
        row   = [rank, "Manual", "", "", title, "", "", added, "", url, ""]
        ws.append(row)
        for col in range(1, len(row) + 1):
            c = ws.cell(ws.max_row, col)
            c.fill = MANUAL_FILL; c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
        if url:
            c = ws.cell(ws.max_row, 10)
            c.hyperlink = url; c.font = URL_FONT

    widths = [4, 8, 6, 18, 42, 28, 20, 10, 14, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    ws2 = wb.create_sheet("Legend")
    ws2.cell(1, 1).value = "Color key"
    ws2.cell(1, 1).font  = Font(bold=True)
    ws2.cell(2, 1).fill  = SCRAPED_FILL
    ws2.cell(2, 2).value = "Auto-scraped & scored job"
    ws2.cell(3, 1).fill  = MANUAL_FILL
    ws2.cell(3, 2).value = "Manually added job URL"

    path = f"/tmp/apply_list_{date_str}.xlsx"
    wb.save(path)
    return path


# ── command handlers ──────────────────────────────────────────────────────────

def do_apply():
    sl      = get_shortlist()
    manual  = get_manual()
    sl_jobs = sl.get("jobs", [])
    m_ents  = manual.get("entries", [])

    if not sl_jobs and not m_ents:
        send_text("Your apply list is empty!\nShortlist some jobs from today's digest or paste job URLs here.")
        return

    send_text(
        f"⏳ Generating apply list…\n"
        f"✅ {len(sl_jobs)} shortlisted  +  🔗 {len(m_ents)} manual"
    )
    path = make_apply_excel(sl_jobs, m_ents, session_date)
    if not path:
        send_text("ERROR: openpyxl not installed — cannot generate Excel.")
        return

    res = send_document(
        path,
        caption=(
            f"📋 Apply List — {session_date}\n"
            f"✅ {len(sl_jobs)} shortlisted  +  🔗 {len(m_ents)} manual\n"
            f"Total: {len(sl_jobs) + len(m_ents)} jobs"
        ),
    )
    if not res.get("ok"):
        send_text(f"WARNING: couldn't send Excel: {res}")

def do_status():
    sl     = get_shortlist()
    manual = get_manual()
    lines  = [
        f"📊 <b>Status — {session_date}</b>",
        f"✅ Shortlisted: <b>{len(sl.get('jobs', []))}</b> jobs",
        f"🔗 Manual links: <b>{len(manual.get('entries', []))}</b>",
        "",
        "Send /apply to get your final Excel.",
    ]
    if sl.get("jobs"):
        lines.append("\n<b>Your shortlist:</b>")
        for i, j in enumerate(sl["jobs"], 1):
            lines.append(f"{i}. {j.get('title','')} @ {j.get('company','')}")
    send_text("\n".join(lines))

def do_clear():
    save_json(SHORTLIST, {"date": session_date, "jobs": []})
    save_json(MANUAL,    {"date": session_date, "entries": []})
    send_text("🗑 Shortlist and manual links cleared for today.")
    return True  # signals a state change


# ── fetch updates ─────────────────────────────────────────────────────────────

last_offset = 0
try:
    last_offset = int(Path(OFFSET_FILE).read_text().strip())
except Exception:
    pass

res     = tg_api("getUpdates", {"offset": last_offset, "limit": 100, "timeout": 0})
updates = res.get("result", [])

debug_lines = [
    f"run_at={datetime.utcnow().isoformat()}",
    f"offset_in={last_offset}",
    f"updates_count={len(updates)}",
    f"jobs_loaded={len(jobs_by_idx)}",
    f"session_date={session_date}",
]

if not updates:
    print("No new updates.")
    debug_lines.append("result=no_updates")
    with open("data/bot_debug.txt", "w") as f:
        f.write("\n".join(debug_lines) + "\n")
    sys.exit(0)

state_changed = False

for update in updates:
    uid          = update["update_id"]
    last_offset  = max(last_offset, uid + 1)

    # ── inline button press ───────────────────────────────────────────────────
    if "callback_query" in update:
        cq     = update["callback_query"]
        cq_id  = cq["id"]
        data   = cq.get("data", "")
        msg_id = cq.get("message", {}).get("message_id")
        debug_lines.append(f"callback uid={uid} data={data!r}")

        if not data.startswith(("sl:", "sk:")):
            answer_callback(cq_id)
            continue

        action, idx_str = data.split(":", 1)
        try:
            idx = int(idx_str)
        except ValueError:
            answer_callback(cq_id, "Invalid callback data.", alert=True)
            debug_lines.append(f"  bad_idx={idx_str!r}")
            continue

        job = jobs_by_idx.get(idx)
        debug_lines.append(f"  idx={idx} job_found={job is not None}")
        if not job:
            answer_callback(cq_id, "Session expired — check tomorrow's digest.", alert=True)
            continue

        if action == "sl":
            sl = get_shortlist()
            already = any(j.get("idx") == idx for j in sl["jobs"])
            debug_lines.append(f"  already_in_list={already}")
            if not already:
                sl["jobs"].append(job)
                save_json(SHORTLIST, sl)
                state_changed = True
                debug_lines.append(f"  saved_shortlist_count={len(sl['jobs'])}")

            answer_callback(cq_id, "✅ Added to shortlist!")

            # Update message text to reflect shortlisted state
            ROLE_EMOJI = {"Product Manager":"🧩","Product Owner":"📋",
                          "Business Analyst":"📊","Consultant/Strategy":"💼","Program Manager":"🗂"}
            sc    = job["score"]
            dot   = "🟢" if sc >= 75 else ("🟡" if sc >= 60 else "🔴")
            emoji = ROLE_EMOJI.get(job.get("role",""), "🔹")
            new_text = (
                f"✅ <b>SHORTLISTED</b> {dot} #{idx+1} · {sc}/100 {emoji} <b>{job['title']}</b>\n"
                f"🏢 {job['company']}"
            )
            if job.get("location"):
                new_text += f" · 📍 {job['location']}"
            if job.get("url"):
                new_text += f'\n<a href="{job["url"]}">View →</a>'
            if msg_id:
                edit_text(msg_id, new_text)

        elif action == "sk":
            answer_callback(cq_id, "❌ Skipped")
            if msg_id:
                remove_keyboard(msg_id)

    # ── text / URL message ────────────────────────────────────────────────────
    elif "message" in update:
        msg      = update["message"]
        text     = (msg.get("text") or "").strip()
        chat_id  = str(msg.get("chat", {}).get("id", ""))
        debug_lines.append(f"message uid={uid} text={text[:40]!r} chat={chat_id}")

        if not text or chat_id != str(CHAT_ID):
            continue

        if text.startswith("/apply"):
            do_apply()

        elif text.startswith("/status"):
            do_status()

        elif text.startswith("/clear"):
            if do_clear():
                state_changed = True

        elif text.startswith(("/help", "/start")):
            send_text(
                "👋 <b>Job Search Bot</b>\n\n"
                "<b>Morning digest:</b> interactive cards arrive daily at 5 AM IST\n"
                "• Tap <b>✅ Shortlist</b> / <b>❌ Skip</b> on each job card\n\n"
                "<b>Manual jobs:</b> paste any job URL here → auto-added\n\n"
                "<b>Commands:</b>\n"
                "• /apply — generate your final apply-list Excel\n"
                "• /status — show shortlist & manual counts\n"
                "• /clear — reset today's list"
            )

        elif text.startswith("http://") or text.startswith("https://"):
            manual = get_manual()
            url    = text.split()[0]  # take first token in case of trailing text
            notes  = text[len(url):].strip()
            already = any(e.get("url") == url for e in manual["entries"])
            if not already:
                manual["entries"].append({
                    "url":      url,
                    "title":    notes or "",
                    "added_at": datetime.utcnow().isoformat(),
                })
                save_json(MANUAL, manual)
                state_changed = True
            count = len(manual["entries"])
            msg_text = f"📌 Added! ({count} manual job{'s' if count != 1 else ''} so far)"
            if notes:
                msg_text += f"\n📝 Note: {notes}"
            msg_text += "\n\nSend /apply when ready for your Excel."
            send_text(msg_text)

# ── save offset & debug log ───────────────────────────────────────────────────

Path(OFFSET_FILE).write_text(str(last_offset))
debug_lines.append(f"offset_out={last_offset}")
debug_lines.append(f"state_changed={state_changed}")
with open("data/bot_debug.txt", "w") as f:
    f.write("\n".join(debug_lines) + "\n")

print(f"Processed {len(updates)} update(s). Offset: {last_offset}. Changed: {state_changed}")
sys.exit(0)
