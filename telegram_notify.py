"""
Daily job digest: sends full Excel + interactive job cards to Telegram.
Called by GitHub Actions after jobspy_run.py.
"""
import json, os, sys, uuid, urllib.request, urllib.error
from datetime import date, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

TOKEN     = os.environ.get("TELEGRAM_TOKEN", "")
CHAT_ID   = os.environ.get("TELEGRAM_CHAT_ID", "")
JSON_PATH = os.environ.get("JOBS_JSON", "jobs_raw.json")

if not TOKEN or not CHAT_ID:
    print("WARNING: TELEGRAM_TOKEN or TELEGRAM_CHAT_ID not set — skipping.")
    sys.exit(0)


# ── scoring ───────────────────────────────────────────────────────────────────

def role_score(t):
    t = t.lower()
    for kw in ["principal product manager", "senior product manager",
                "lead product manager", "staff product manager", "product manager, senior"]:
        if kw in t: return 40, "Product Manager"
    if "product manager" in t:   return 40, "Product Manager"
    if "product owner" in t:     return 37, "Product Owner"
    for kw in ["senior business analyst", "lead business analyst", "technical business analyst"]:
        if kw in t: return 35, "Business Analyst"
    if "business analyst" in t:  return 35, "Business Analyst"
    for kw in ["consultant, strategy", "strategy, growth", "transformation consultant",
                "management consultant", "digital transformation", "strategy consultant"]:
        if kw in t: return 33, "Consultant/Strategy"
    if "consultant" in t:        return 33, "Consultant/Strategy"
    for kw in ["program manager", "delivery manager", "project manager"]:
        if kw in t: return 28, "Program Manager"
    for kw in ["product specialist", "product analyst", "product associate", "product operations"]:
        if kw in t: return 28, "Product Manager"
    return 0, None

def skills_pts(d):
    if not d: return 0
    d = d.lower(); p = 0
    if "roadmap" in d or "product strategy" in d:          p += 5
    if any(x in d for x in ["backlog", "sprint", "agile", "scrum"]): p += 4
    if any(x in d for x in ["requirements", "acceptance criteria", "uat"]): p += 4
    if "stakeholder" in d:                                  p += 4
    if any(x in d for x in ["sql", "power bi", "kpi", "analytics"]): p += 4
    if any(x in d for x in ["generative ai", "machine learning", "llm",
                              "agentic", "ai/ml", "ai product", "ai-powered"]): p += 5
    if any(x in d for x in ["digital transformation", "process improvement", "automation"]): p += 4
    return min(p, 30)

def domain_pts(d):
    if not d: return 0
    d = d.lower(); p = 0
    if any(x in d for x in ["insurance", "motor claims", "claims"]):     p += 10
    elif any(x in d for x in ["bfsi", "fintech", "financial services", "banking", "payments"]): p += 7
    elif any(x in d for x in ["regtech", "risk management", "compliance"]): p += 4
    if any(x in d for x in ["mba or relevant advanced degree is a must",
                              "mba or another advanced degree", "mba preferred"]): p += 5
    if any(x in d for x in ["ai product", "generative ai", "agentic", "ai platform", "ai strategy"]): p += 5
    return min(p, 20)

def brand_pts(c, d=""):
    c = (c or "").lower(); d = (d or "").lower()
    if any(x in c for x in ["amazon", "google", "microsoft", "salesforce", "jpmorgan",
                              "goldman", "wells fargo", "blackrock"]): return 10
    if "hackajob" in c and "jp morgan" in d: return 10
    if any(x in c for x in ["deloitte", "pwc", "hsbc", "synchrony", "nielsen", "experian",
                              "optum", "simcorp", "natwest", "state street", "broadridge", "wise"]): return 7
    if any(x in c for x in ["wipro", "infosys", "tcs", "genpact", "capgemini", "accenture",
                              "cognizant", "publicis sapient", "endava"]): return 4
    return 0

def salary_str(j):
    lo = j.get("min_amount"); hi = j.get("max_amount"); cur = j.get("currency") or ""
    if lo and hi:  return f"{cur}{int(lo):,}–{int(hi):,}"
    if lo:         return f"{cur}{int(lo):,}+"
    return ""


# ── load & score ──────────────────────────────────────────────────────────────

try:
    with open(JSON_PATH) as f:
        raw = json.load(f)
except Exception as e:
    print(f"ERROR reading {JSON_PATH}: {e}")
    sys.exit(0)

run_date_str = raw.get("run_date", str(date.today()))
try:
    run_date = date.fromisoformat(run_date_str)
except ValueError:
    run_date = date.today()
cutoff = (run_date - timedelta(hours=48)).isoformat()

scored = []
seen   = set()

for j in raw.get("jobs", []):
    dp      = j.get("date_posted")
    title   = j.get("title", "") or ""
    company = j.get("company", "") or ""
    desc    = j.get("description", "") or ""
    if dp and dp < cutoff:
        continue
    r1, rt = role_score(title)
    if r1 == 0:
        continue
    key = (company.lower()[:20], title.lower()[:35])
    if key in seen:
        continue
    seen.add(key)
    total = r1 + skills_pts(desc) + domain_pts(desc) + brand_pts(company, desc)
    if total < 55:
        continue
    scored.append({
        "idx":      len(scored),
        "score":    total,
        "title":    title,
        "company":  company,
        "url":      j.get("job_url", "") or "",
        "date":     dp or "Recent",
        "role":     rt or "",
        "location": j.get("location", "") or "",
        "salary":   salary_str(j),
        "site":     j.get("site", "") or "",
        "level":    j.get("job_level", "") or "",
    })

scored.sort(key=lambda x: -x["score"])
for i, j in enumerate(scored):
    j["idx"] = i


# ── save state files ──────────────────────────────────────────────────────────

Path("data").mkdir(exist_ok=True)

with open("data/today_jobs.json", "w") as f:
    json.dump({"date": run_date_str, "jobs": scored}, f, indent=2)

for path, template in [
    ("data/shortlisted.json", {"date": run_date_str, "jobs": []}),
    ("data/manual_jobs.json", {"date": run_date_str, "entries": []}),
]:
    p = Path(path)
    existing = {}
    if p.exists():
        try: existing = json.loads(p.read_text())
        except: pass
    if existing.get("date") != run_date_str:
        with open(path, "w") as f:
            json.dump(template, f, indent=2)


# ── Excel generation ──────────────────────────────────────────────────────────

def make_full_excel(jobs, date_str):
    if not HAS_OPENPYXL:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Scored Jobs"

    HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
    GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
    YELLOW_FILL= PatternFill("solid", fgColor="FFEB9C")
    RED_FILL   = PatternFill("solid", fgColor="FFC7CE")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    BD         = Side(style="thin", color="CCCCCC")
    BORDER     = Border(left=BD, right=BD, top=BD, bottom=BD)

    cols = ["#", "Score", "Role", "Title", "Company", "Location", "Date", "Salary", "Site", "Apply URL"]
    ws.append(cols)
    for col in range(1, len(cols) + 1):
        c = ws.cell(1, col)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER

    for rank, j in enumerate(jobs, 1):
        sc = j["score"]
        row = [rank, sc, j["role"], j["title"], j["company"],
               j["location"], j["date"], j["salary"], j["site"], j["url"]]
        ws.append(row)
        fill = GREEN_FILL if sc >= 75 else (YELLOW_FILL if sc >= 60 else RED_FILL)
        for col in range(1, len(row) + 1):
            c = ws.cell(ws.max_row, col)
            c.fill = fill; c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
        if j["url"]:
            c = ws.cell(ws.max_row, 10)
            c.hyperlink = j["url"]
            c.font = Font(color="0563C1", underline="single")

    widths = [4, 6, 18, 42, 28, 20, 10, 14, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    # Legend
    ws2 = wb.create_sheet("Legend")
    for row, (color, meaning) in enumerate([
        ("E2EFDA", "High match (score ≥ 75)"),
        ("FFEB9C", "Medium match (score 60–74)"),
        ("FFC7CE", "Lower match (score 55–59)"),
    ], start=1):
        ws2.cell(row, 1).fill = PatternFill("solid", fgColor=color)
        ws2.cell(row, 1).value = f"Score {'≥75' if row==1 else ('60–74' if row==2 else '55–59')}"
        ws2.cell(row, 2).value = meaning

    path = f"/tmp/jobs_full_{date_str}.xlsx"
    wb.save(path)
    return path


# ── Telegram helpers ──────────────────────────────────────────────────────────

def tg_api(method, payload):
    url  = f"https://api.telegram.org/bot{TOKEN}/{method}"
    data = json.dumps(payload).encode()
    req  = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        print(f"WARNING: {method} HTTP {e.code}: {e.read().decode(errors='replace')}")
        return {}
    except Exception as e:
        print(f"WARNING: {method} failed: {e}")
        return {}

def send_text(text):
    return tg_api("sendMessage", {
        "chat_id": CHAT_ID, "text": text,
        "parse_mode": "HTML", "disable_web_page_preview": True,
    })

def send_document(path, caption=""):
    boundary = uuid.uuid4().hex
    filename = os.path.basename(path)
    with open(path, "rb") as f:
        file_bytes = f.read()
    body = (
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"chat_id\"\r\n\r\n{CHAT_ID}\r\n"
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"caption\"\r\n\r\n{caption}\r\n"
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"document\"; filename=\"{filename}\"\r\n"
        f"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
    ).encode() + file_bytes + f"\r\n--{boundary}--\r\n".encode()
    url = f"https://api.telegram.org/bot{TOKEN}/sendDocument"
    req = urllib.request.Request(url, data=body,
                                  headers={"Content-Type": f"multipart/form-data; boundary={boundary}"})
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        print(f"WARNING: sendDocument HTTP {e.code}: {e.read().decode(errors='replace')}")
        return {}
    except Exception as e:
        print(f"WARNING: sendDocument failed: {e}")
        return {}


# ── send Excel ────────────────────────────────────────────────────────────────

excel_path = make_full_excel(scored, run_date_str)
if excel_path:
    res = send_document(excel_path,
                        caption=f"📊 Full scored jobs — {run_date_str} ({len(scored)} roles, score ≥55)")
    if res.get("ok"):
        print(f"✅ Excel sent ({len(scored)} jobs)")
    else:
        print(f"WARNING: Excel send result: {res}")
else:
    print("openpyxl not available — skipping Excel")


# ── send interactive job cards (top 20) ──────────────────────────────────────

ROLE_EMOJI = {
    "Product Manager":     "🧩",
    "Product Owner":       "📋",
    "Business Analyst":    "📊",
    "Consultant/Strategy": "💼",
    "Program Manager":     "🗂",
}

def score_dot(s):
    return "🟢" if s >= 75 else ("🟡" if s >= 60 else "🔴")

top = scored[:20]

send_text(
    f"<b>📌 Job Digest — {run_date_str}</b>\n"
    f"<i>{len(scored)} qualifying roles · Tap ✅ to shortlist, ❌ to skip</i>"
)

sent_msgs = {}
for job in top:
    idx   = job["idx"]
    sc    = job["score"]
    emoji = ROLE_EMOJI.get(job["role"], "🔹")

    text = (
        f"{score_dot(sc)} <b>#{idx+1} · {sc}/100</b> {emoji} <b>{job['title']}</b>\n"
        f"🏢 {job['company']}"
    )
    if job["location"]:
        text += f" · 📍 {job['location']}"
    text += f" · 📅 {job['date']}"
    if job["salary"]:
        text += f" · 💰 {job['salary']}"
    if job["url"]:
        text += f'\n<a href="{job["url"]}">View →</a>'

    r = tg_api("sendMessage", {
        "chat_id":    CHAT_ID,
        "text":       text,
        "parse_mode": "HTML",
        "disable_web_page_preview": True,
        "reply_markup": {"inline_keyboard": [[
            {"text": "✅ Shortlist", "callback_data": f"sl:{idx}"},
            {"text": "❌ Skip",      "callback_data": f"sk:{idx}"},
        ]]},
    })
    if r.get("ok") and r.get("result"):
        sent_msgs[str(idx)] = r["result"]["message_id"]

send_text(
    "━━━━━━━━━━━━━━━━━━━━\n"
    "💡 <b>How to use:</b>\n"
    "• Tap ✅ <b>Shortlist</b> to save a job\n"
    "• Paste any job URL to add it manually\n"
    "• /apply — get your final apply-list Excel\n"
    "• /status — see your shortlist count"
)

with open("data/sent_messages.json", "w") as f:
    json.dump({"date": run_date_str, "messages": sent_msgs}, f, indent=2)

print(f"Done. {len(top)} interactive cards sent for {run_date_str}.")
sys.exit(0)
